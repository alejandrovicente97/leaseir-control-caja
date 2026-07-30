# -*- coding: utf-8 -*-
"""
============================================================================
 LEASEIR - Capa de fuentes de datos
============================================================================
 Normaliza a un esquema unico los datos vengan de donde vengan:
   - holded.json   (lo que deja el extractor que corre en el PC)
   - los Excel de la carpeta 19. Control Caja  (modo respaldo / validacion)

 Esquema normalizado
 -------------------
 ventas      : num, cliente, cuenta, fecha, vencimiento, total, cobrado,
               pendiente, estado, fecha_cobro, mes_venc
 compras     : num, proveedor, cuenta, fecha, vencimiento, total, pagado,
               pendiente, estado, fecha_pago, mes_venc, tipologia
 calendario  : factura, cliente, mes, importe        (vencimientos de Eli)
 bancos      : cuenta, saldo, tipo ('cuenta' | 'poliza'), limite
============================================================================
"""
from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime, date
from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
# utilidades
# ---------------------------------------------------------------------------
def norm(txt) -> str:
    """Normaliza texto para comparar: sin tildes, sin dobles espacios, mayusculas."""
    if txt is None:
        return ""
    s = str(txt).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).upper()


def num(v) -> float:
    """Convierte a float tolerando formatos contables ('  -   EUR', '1.234,56')."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return 0.0 if pd.isna(v) else float(v)
    s = str(v).replace("€", "").replace("EUR", "").strip()
    if s in ("", "-", "--"):
        return 0.0
    s = s.replace(".", "").replace(",", ".") if re.search(r",\d{1,2}$", s) else s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def a_fecha(v):
    # pd.NaT es subclase de datetime y su .year es NaN: hay que cazarlo antes
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):          # timestamp unix de Holded
        try:
            return datetime.fromtimestamp(float(v)).date()
        except (OSError, ValueError):
            return None
    s = str(v).strip()[:19]
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
              "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


def mes_de(d) -> str | None:
    d = a_fecha(d)
    if not isinstance(d, (date, datetime)):
        return None
    return f"{d.year}{d.month:02d}"


def suma_meses(mes: str, n: int) -> str:
    a, m = int(mes[:4]), int(mes[4:])
    t = (a * 12 + m - 1) + n
    return f"{t // 12}{t % 12 + 1:02d}"


def nombre_mes(mes: str) -> str:
    MES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
           "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return f"{MES[int(mes[4:]) - 1]} {mes[:4]}"


# ===========================================================================
#  FUENTE 1 - HOLDED JSON  (lo que produce holded_extract.py)
# ===========================================================================
# Esquema canonico. Se declara explicitamente para que un bloque vacio de
# Holded produzca un DataFrame vacio PERO CON COLUMNAS. Sin esto, si un
# endpoint falla el motor casca con KeyError en lugar de dar cero.
COLS_DOC = ["id", "num", "tercero", "cuenta", "fecha", "vencimiento", "total",
            "liquidado", "pendiente", "estado", "estado_api", "fecha_liq",
            "mes_venc", "mes_factura", "tipologia"]
COLS_REAL = ["fecha", "mes", "sentido", "tercero", "num", "doc_id", "importe",
             "banco", "concepto", "conciliado", "tipo_api"]
COLS_PLAN = ["numero", "nombre", "grupo", "debe", "haber", "saldo"]
COLS_DIARIO = ["fecha", "mes", "asiento", "linea", "cuenta", "cuenta_nombre",
               "grupo_pgc", "concepto", "doc", "tipo", "debe", "haber", "importe"]


# Grupos del Plan General Contable, por el primer digito o los tres primeros.
# Sirve para que un movimiento de banco deje de ser un texto y tenga naturaleza:
# "EMISION REMESA SEPA SDD 0049" contra la 430 es un cobro de clientes.
GRUPOS_PGC = [
    ("400", "Proveedores"), ("401", "Proveedores, efectos"),
    ("410", "Acreedores"),  ("411", "Acreedores, efectos"),
    ("430", "Clientes"),    ("431", "Clientes, efectos"),
    ("432", "Clientes, factoring"), ("436", "Clientes de dudoso cobro"),
    ("440", "Deudores"),    ("465", "Remuneraciones pendientes"),
    ("460", "Anticipos de remuneraciones"),
    ("470", "Hacienda deudora"), ("475", "Hacienda acreedora"),
    ("476", "Seguridad Social acreedora"), ("471", "Seguridad Social deudora"),
    ("520", "Deudas a corto con entidades de credito"),
    ("523", "Proveedores de inmovilizado a corto"),
    ("527", "Intereses de deudas a corto"),
    ("170", "Deudas a largo con entidades de credito"),
    ("572", "Bancos"), ("570", "Caja"), ("574", "Bancos"),
    ("640", "Sueldos y salarios"), ("642", "Seguridad Social a cargo empresa"),
    ("621", "Arrendamientos"), ("623", "Servicios profesionales"),
    ("624", "Transportes"),
    ("662", "Intereses de deudas"), ("665", "Intereses por descuento de efectos"),
    ("669", "Otros gastos financieros"), ("668", "Diferencias de cambio"),
    ("66",  "Gastos financieros"),
    ("551", "Cuenta corriente con socios y administradores"),
    ("552", "Cuenta corriente con empresas del grupo"),
    ("553", "Cuenta corriente con socios"),
    # 555 NO es financiacion: son apuntes que aun no se han aplicado a su
    # cuenta definitiva. Meterlos en el puente del unlevered cambiaba la cifra
    # de -2.449 a -133.569, que es toda la diferencia del mundo.
    ("555", "Partidas pendientes de aplicacion"),
    ("556", "Desembolsos exigidos"),
    ("57",  "Tesoreria"),
    ("41",  "Acreedores varios"), ("44", "Deudores varios"),
    ("47",  "Administraciones publicas"),
    ("52",  "Deudas a corto plazo"), ("17", "Deudas a largo plazo"),
    ("60",  "Compras"), ("62", "Servicios exteriores"), ("63", "Tributos"),
    ("64",  "Gastos de personal"), ("70", "Ventas"),
]


def grupo_pgc(cuenta) -> str:
    c = re.sub(r"\D", "", str(cuenta or ""))
    for pref, nombre in GRUPOS_PGC:
        if c.startswith(pref):
            return nombre
    return f"Grupo {c[:1]}" if c else "Sin cuenta"
COLS_BANCO = ["cuenta", "saldo", "tipo", "limite"]
COLS_MOV = ["cuenta", "fecha", "importe", "concepto"]


def clasificar_cuenta(nombre: str, tipo_api: str = "") -> str:
    """
    cuenta | poliza | tarjeta, a partir del nombre.

    Holded devuelve type="bank" para todo, asi que el nombre es lo unico que
    distingue. Y el orden importa: "TARJETA NEGOCIOS CREDITO" lleva CREDITO
    dentro, pero es una tarjeta, no una poliza. Si se mira primero el credito,
    las tarjetas acaban sumando como disponible de financiacion, que es
    justo lo contrario de lo que son.
    """
    n = norm(nombre).lower()
    # "T.CAIXA ARTURO 5139", "T. Santander Maria Magaz" y "Santander VIA T"
    # tambien son tarjetas: abreviadas y de peaje. Sin esto se colaban en la
    # caja como cuenta corriente, con su saldo negativo incluido.
    if (any(p in n for p in ("tarjeta", "card", "visa", "mastercard", "amex",
                             "american express", "via t"))
            or re.match(r"^t\.?\s?[a-z]*\s?\d*$", n)
            or re.match(r"^t\.", n)):
        return "tarjeta"
    # "Santander Cred. (7628)" es una poliza abreviada
    if ("poliz" in n or "credit" in n or "cred." in n
            or "linea de credito" in n or "credit" in (tipo_api or "").lower()):
        return "poliza"
    return "cuenta"


def _clave(k: str) -> str:
    """paymentsPending, payments_pending y PaymentsPending son la misma cosa."""
    return re.sub(r"[^a-z0-9]", "", str(k).lower())


def _pri(d: dict, *claves, defecto=None):
    """
    Primer valor no vacio de una lista de claves candidatas, comparando sin
    distinguir mayusculas ni guiones bajos.

    Hace falta porque la API v1 de Holded devuelve camelCase (paymentsPending)
    y la v2 snake_case (payments_pending). Sin esto, contra la v2 no se
    encontraba lo cobrado y TODA factura salia como pendiente: el dashboard
    daba 28 millones de pendiente de cobro en vez de un millon.
    """
    idx = getattr(d, "_idx_leaseir", None)
    if idx is None:
        idx = {_clave(k): k for k in d}
        try:
            d._idx_leaseir = idx
        except AttributeError:
            pass
    for c in claves:
        real = idx.get(_clave(c))
        if real is not None and d[real] not in (None, "", []):
            return d[real]
    return defecto


def desde_holded(ruta: Path) -> dict:
    """
    Normaliza el JSON del extractor.

    Los nombres de campo de Holded no son estables entre endpoints ni entre
    versiones, asi que cada dato se busca en varias claves candidatas en vez de
    dar una por segura. Un cambio de nombre en Holded degrada un campo, no
    tumba el forecast.
    """
    with open(ruta, encoding="utf-8") as f:
        d = json.load(f)

    meta = d.get("_meta", {}) or {}
    contactos = {}
    for c in d.get("contactos", []) or []:
        if isinstance(c, dict):
            cid = _pri(c, "id", "_id")
            if cid:
                contactos[cid] = _pri(c, "name", "tradeName", "legalName",
                                      "socialName", defecto="")

    def documentos(clave, signo=1):
        filas = []
        for doc in d.get(clave, []) or []:
            if not isinstance(doc, dict):
                continue
            total = num(_pri(doc, "total", "totalAmount", "totalWithTax", "amount",
                             "grandTotal", defecto=0)) * signo
            cobrado = num(_pri(doc, "paymentsTotal", "paidAmount", "paid",
                               "amountPaid", "collectedAmount", defecto=0)) * signo
            # Holded suele dar el pendiente ya calculado: es mas fiable que restar
            pend_api = _pri(doc, "paymentsPending", "pendingAmount", "pending",
                            "amountPending", "outstandingAmount", "dueAmount")
            pendiente = num(pend_api) * signo if pend_api is not None else total - cobrado

            venc = a_fecha(_pri(doc, "dueDate", "expirationDate", "date", "issueDate"))
            emision = a_fecha(_pri(doc, "date", "issueDate", "issuedDate",
                                   "documentDate", "createdAt"))
            f_liq = a_fecha(_pri(doc, "paymentDate", "paidDate", "lastPaymentDate",
                                 "collectionDate", "settledAt"))
            # si esta cobrada del todo y no hay fecha de cobro, vale el vencimiento
            if f_liq is None and abs(pendiente) < 0.01 and abs(total) > 0.01:
                f_liq = venc

            estado_api = _pri(doc, "status", "state")
            if abs(pendiente) < 0.01 and abs(total) > 0.01:
                estado = "Pagado"
            elif venc and venc < date.today():
                estado = "Vencido"
            else:
                estado = "Pendiente"

            tags = _pri(doc, "tags", defecto=[]) or []
            tag = tags[0] if isinstance(tags, list) and tags else (
                tags if isinstance(tags, str) else "")

            filas.append({
                # id interno de Holded: es la clave por la que los apuntes de
                # /payments enganchan con su factura (document_id). Sin el, el
                # detalle de cobros y pagos realizados no se puede desglosar
                # hasta factura, solo hasta tercero.
                "id":          str(_pri(doc, "id", "_id", defecto="")),
                "num":         _pri(doc, "docNumber", "documentNumber", "invoiceNum",
                                    "number", "id", defecto=""),
                "tercero":     _pri(doc, "contactName", "contactLegalName", "clientName",
                                    "supplierName",
                                    defecto=contactos.get(
                                        _pri(doc, "contact", "contactId"), "")),
                "cuenta":      str(_pri(doc, "desc", "description", "notes",
                                        "concept", defecto=""))[:90],
                "fecha":       emision,
                "vencimiento": venc,
                "total":       total,
                "liquidado":   cobrado,
                "pendiente":   pendiente,
                "estado":      estado,
                "estado_api":  estado_api,
                "fecha_liq":   f_liq,
                "mes_venc":    mes_de(venc),
                "mes_factura": mes_de(emision),
                "tipologia":   tag,
            })
        return filas

    ventas = documentos("facturas_venta") + documentos("recibos_venta") \
        + documentos("abonos_venta", -1)
    compras = documentos("facturas_compra") + documentos("abonos_compra", -1)

    # ---- posicion bancaria ------------------------------------------------
    bancos = []
    for c in d.get("cuentas_tesoreria", []) or []:
        if not isinstance(c, dict):
            continue
        if _pri(c, "archived", defecto=False) in (True, "true", 1):
            continue
        nombre = _pri(c, "name", "alias", "description", "accountName",
                      defecto="(sin nombre)")
        tipo_api = str(_pri(c, "type", "accountType", "kind", defecto="")).lower()
        bancos.append({
            "cuenta": nombre,
            "saldo": num(_pri(c, "balance", "currentBalance", "currentAmount",
                              "amount", defecto=0)),
            "tipo": clasificar_cuenta(nombre, tipo_api),
            # Holded NO publica el limite de la poliza: en el esquema de
            # treasury/accounts no hay ningun campo de limite de credito. Sale
            # de config.yaml, escrito a mano. Cero aqui significa "sin
            # configurar", no "sin disponible": son cosas muy distintas y el
            # dashboard las distingue.
            "limite": 0.0,
        })

    # ---- cobros y pagos REALIZADOS ---------------------------------------
    # /payments es el libro de liquidaciones: cada apunte trae el documento al
    # que va (document_id), la fecha, el importe y la cuenta. Es la unica forma
    # de dar el ejecutado del mes desglosado hasta factura sin inventarselo a
    # partir de la fecha de liquidacion de la factura, que solo guarda la del
    # ultimo pago y se lleva por delante los cobros parciales.
    idx_venta = {f["id"]: f for f in ventas if f.get("id")}
    idx_compra = {f["id"]: f for f in compras if f.get("id")}
    cuentas_nom = {str(c.get("id") or c.get("_id")): c.get("name")
                   for c in (d.get("cuentas_tesoreria") or [])
                   if isinstance(c, dict)}

    realizados = []
    for p in d.get("pagos", []) or []:
        if not isinstance(p, dict):
            continue
        doc_id = str(_pri(p, "document_id", "documentId", "docId",
                          "invoice_id", defecto="") or "")
        fecha = a_fecha(_pri(p, "date", "paymentDate", "valueDate", "createdAt"))
        importe = num(_pri(p, "amount", "total", "value", defecto=0))
        if not fecha or abs(importe) < 0.005:
            continue

        # El sentido se decide por el documento al que apunta, no por el
        # vocabulario de document_type: asi no depende de como Holded llame
        # hoy a cada tipo de documento.
        fac = idx_venta.get(doc_id)
        if fac is not None:
            sentido = "cobro"
        elif doc_id in idx_compra:
            fac = idx_compra[doc_id]
            sentido = "pago"
        else:
            # Sin documento que lo respalde NO se adivina el sentido. Holded
            # manda el importe siempre en positivo, asi que dar por cobro todo
            # lo que no case metia en "cobros realizados" las cuotas de
            # tarjeta, los peajes, los impuestos y las transferencias a
            # proveedores. Alejandro lo vio de un vistazo: "eso que marcas como
            # cobros son pagos".
            # Un apunte mal firmado es peor que un apunte que falta: el que
            # falta se busca, el mal firmado se cree. Van a su propio cajon.
            fac = None
            sentido = "sin_documento"

        realizados.append({
            "fecha": fecha,
            "mes": mes_de(fecha),
            "sentido": sentido,
            "tercero": (fac or {}).get("tercero") or _pri(
                p, "contact_name", "contactName", defecto="(sin tercero)"),
            "num": (fac or {}).get("num") or "(sin factura)",
            "doc_id": doc_id,
            # signo de caja: los cobros entran, los pagos salen. Sin documento
            # no hay signo fiable, se deja el valor absoluto tal cual y el
            # panel lo muestra aparte sin sumarlo a ningun total.
            "importe": (abs(importe) if sentido == "cobro"
                        else -abs(importe) if sentido == "pago" else importe),
            "tipo_api": str(_pri(p, "type", "document_type", "documentType",
                                 defecto="")),
            "banco": cuentas_nom.get(str(_pri(p, "bank_account_id", "bankAccountId",
                                              "treasuryId", defecto="")), ""),
            "concepto": str(_pri(p, "description", "desc", "notes", defecto=""))[:90],
            "conciliado": _pri(p, "reconciliation_status", "reconciliationStatus",
                               "status", defecto=""),
        })

    # ---- plan contable ----------------------------------------------------
    plan = []
    for c in d.get("plan_contable", []) or []:
        if not isinstance(c, dict):
            continue
        plan.append({
            "numero": str(_pri(c, "number", "num", "code", defecto="")),
            "nombre": _pri(c, "name", "description", defecto=""),
            "grupo":  _pri(c, "group", defecto=""),
            # Holded manda los importes como cadena decimal, a proposito, para
            # no perder precision. num() los convierte respetando el formato.
            "debe":   num(_pri(c, "debit", "debe", defecto=0)),
            "haber":  num(_pri(c, "credit", "haber", defecto=0)),
            "saldo":  num(_pri(c, "balance", "saldo", defecto=0)),
        })

    # ---- libro diario -----------------------------------------------------
    # nombre real de cada cuenta, del plan contable. Sin esto el puente dice
    # "52000042 Deudas a corto con entidades de credito" cuando en realidad es
    # la tarjeta American Express, y no hay forma de discutir el criterio.
    nom_cta = {str(c["numero"]): c["nombre"] for c in plan if c.get("numero")}
    diario = []
    for e in d.get("libro_diario", []) or []:
        if not isinstance(e, dict):
            continue
        f = a_fecha(_pri(e, "date", "entryDate", "fecha"))
        debe = num(_pri(e, "debit", "debe", defecto=0))
        haber = num(_pri(e, "credit", "haber", defecto=0))
        cta = _pri(e, "account", "accountNumber", "cuenta", defecto="")
        diario.append({
            "fecha": f, "mes": mes_de(f),
            "asiento": _pri(e, "entry_number", "entryNumber", defecto=""),
            "linea": _pri(e, "line", defecto=""),
            "cuenta": str(cta),
            "cuenta_nombre": nom_cta.get(str(cta), ""),
            "grupo_pgc": grupo_pgc(cta),
            "concepto": str(_pri(e, "description", "desc", defecto=""))[:90],
            "doc": str(_pri(e, "doc_description", "docDescription", defecto=""))[:90],
            "tipo": str(_pri(e, "type", "tipo", defecto="") or ""),
            "debe": debe, "haber": haber,
            # signo de caja: en una cuenta de banco el debe entra y el haber sale
            "importe": debe - haber,
        })

    movs = []
    for m in d.get("movimientos_tesoreria", []) or []:
        if not isinstance(m, dict):
            continue
        movs.append({
            "cuenta": m.get("_cuenta_nombre"),
            "fecha": a_fecha(_pri(m, "date", "valueDate", "operationDate", "createdAt")),
            "importe": num(_pri(m, "amount", "value", "total", defecto=0)),
            "concepto": _pri(m, "description", "concept", "desc", "notes", defecto=""),
        })

    sello = meta.get("extraido_en", "?")
    return {
        "ventas": pd.DataFrame(ventas, columns=COLS_DOC).rename(
            columns={"tercero": "cliente"}),
        "compras": pd.DataFrame(compras, columns=COLS_DOC).rename(
            columns={"tercero": "proveedor"}),
        "bancos": pd.DataFrame(bancos, columns=COLS_BANCO),
        "movimientos": pd.DataFrame(movs, columns=COLS_MOV),
        "realizados": pd.DataFrame(realizados, columns=COLS_REAL),
        "plan_contable": pd.DataFrame(plan, columns=COLS_PLAN),
        "diario": pd.DataFrame(diario, columns=COLS_DIARIO),
        "origen": f"API de Holded ({sello})",
        "avisos_origen": meta.get("avisos", []),
    }


# ===========================================================================
#  FUENTE 2 - EXCEL de la carpeta (respaldo y validacion)
# ===========================================================================
def _hoja(ruta: Path, hoja: str, fila_cabecera: int) -> pd.DataFrame:
    df = pd.read_excel(ruta, sheet_name=hoja, header=fila_cabecera - 1, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def desde_excel(f_cobros: Path, f_forecast: Path) -> dict:
    # -- facturas de venta (pestana Holded del control de cobros) -----------
    v = _hoja(f_cobros, "Holded", 5)
    ventas = pd.DataFrame({
        "num":         v["Num"],
        "cliente":     v["Cliente"],
        "cuenta":      v["Cuenta"],
        "fecha":       v["Fecha"].map(a_fecha),
        "vencimiento": v["Vencimiento"].map(a_fecha),
        "total":       v["Total"].map(num),
        "liquidado":   v["Cobrado"].map(num),
        "estado":      v["Estado"],
        "fecha_liq":   v["Fecha de cobro"].map(a_fecha),
    }).dropna(subset=["num"])
    ventas["pendiente"] = ventas["total"] - ventas["liquidado"]
    ventas["mes_venc"] = ventas["vencimiento"].map(mes_de)
    ventas["tipologia"] = ""
    # "Ano Mes" es el mes de emision real; la columna Fecha es la de extraccion
    ventas["mes_factura"] = (v["Año Mes"].map(lambda x: str(int(x)) if isinstance(x, (int, float))
                                              and not pd.isna(x) else (str(x)[:6] if x else None))
                             if "Año Mes" in v.columns else None)

    # -- facturas de compra (pestana Proveedores del forecast) --------------
    c = _hoja(f_forecast, "Proveedores", 5)
    compras = pd.DataFrame({
        "num":         c["Num"],
        "proveedor":   c["Proveedor"],
        "cuenta":      c["Cuenta"],
        "fecha":       c["Fecha emisión"].map(a_fecha),
        "vencimiento": c["Vencimiento"].map(a_fecha),
        "total":       c["Total"].map(num),
        "liquidado":   c["Pagado"].map(num),
        "estado":      c["Estado"],
        "fecha_liq":   c["Fecha de pago"].map(a_fecha),
        "tipologia":   c["Tipología"],
    }).dropna(subset=["proveedor"])
    compras["pendiente"] = compras["total"] - compras["liquidado"]
    compras["mes_venc"] = compras["vencimiento"].map(mes_de)

    # -- posicion bancaria (hoja Forecast Caja - Mes en Curso) --------------
    import openpyxl
    wb = openpyxl.load_workbook(f_forecast, read_only=True, data_only=True)
    ws = wb["Forecast Caja - Mes en Curso"]
    bancos = []
    for fila in range(121, 129):                       # cuentas corrientes
        nom, sal = ws.cell(fila, 4).value, ws.cell(fila, 6).value
        if nom and sal is not None:
            bancos.append({"cuenta": str(nom), "saldo": num(sal), "tipo": "cuenta", "limite": 0.0})
    for fila in range(131, 134):                       # polizas de credito
        nom, sal = ws.cell(fila, 4).value, ws.cell(fila, 6).value
        if nom and sal is not None and num(sal) != 0:
            bancos.append({"cuenta": str(nom), "saldo": num(sal), "tipo": "poliza",
                           "limite": num(ws.cell(fila, 3).value)})
    wb.close()

    return {
        "ventas": ventas,
        "compras": compras,
        "bancos": pd.DataFrame(bancos),
        "movimientos": pd.DataFrame(columns=COLS_MOV),
        # el Excel no lleva libro de liquidaciones: el ejecutado por factura
        # solo existe con la API. Vacio PERO CON COLUMNAS.
        "realizados": pd.DataFrame(columns=COLS_REAL),
        "plan_contable": pd.DataFrame(columns=COLS_PLAN),
        "diario": pd.DataFrame(columns=COLS_DIARIO),
        "origen": f"Excel {f_forecast.name}",
    }


# ===========================================================================
#  CALENDARIO DE COBROS DE ELI  (pestana ELISABET / Google Sheet)
# ===========================================================================
def calendario_cobros(ruta: Path, hoja: str = "ELISABET") -> pd.DataFrame:
    """
    La hoja es una matriz factura x mes: cada celda es la cuota que se espera
    cobrar de esa factura en ese mes. La convertimos a formato largo.
    """
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[hoja]

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return pd.DataFrame(columns=["factura", "cliente", "mes", "importe"])

    cab = filas[0]
    # columnas cuyo encabezado es una fecha -> son los meses del calendario
    meses = {}
    for j, val in enumerate(cab):
        f = a_fecha(val)
        if f and 2020 <= f.year <= 2040:
            meses[j] = f"{f.year}{f.month:02d}"

    registros, presentes = [], set()
    for fila in filas[1:]:
        if len(fila) < 3:
            continue
        cliente, factura = fila[1], fila[2]
        if not factura:
            continue
        # Ojo: una factura puede estar en la hoja SIN calendario (abonos, ventas
        # al contado). Sigue "estando en ELISABET" y hay que distinguirlo de no
        # estar en absoluto: si no, el pendiente de cobro sale negativo.
        presentes.add(norm(factura))
        for j, mes in meses.items():
            if j < len(fila):
                imp = fila[j]
                if isinstance(imp, (int, float)) and imp and abs(imp) > 0.005:
                    registros.append({"factura": str(factura).strip(),
                                      "cliente": str(cliente or "").strip(),
                                      "mes": mes, "importe": float(imp)})
    wb.close()
    df = pd.DataFrame(registros, columns=["factura", "cliente", "mes", "importe"])
    df.attrs["facturas_en_hoja"] = presentes
    return df


def mapping_cuentas(ruta: Path) -> tuple[set, dict]:
    """Devuelve (cuentas que son renting, cuenta -> categoria simple)."""
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Mapping"]
    rentings, simple = set(), {}
    for fila in ws.iter_rows(min_row=3, max_row=300, values_only=True):
        if len(fila) > 2 and fila[2]:
            rentings.add(norm(fila[2]))
        if len(fila) > 6 and fila[5] and fila[6]:
            simple[norm(fila[5])] = str(fila[6]).strip()
    wb.close()
    return rentings, simple


# ===========================================================================
#  CALENDARIO DESDE GOOGLE DRIVE
# ===========================================================================
def calendario_desde_drive(ruta_descargada: Path, hoja: str = "ELISABET") -> pd.DataFrame:
    """
    El fichero de Eli vive en Drive como .xlsx subido
    (https://docs.google.com/spreadsheets/d/1EmO9WHz-ewB8objYRnAvoQ2ZBkhnzAbR).
    Una vez descargado con el conector de Google Drive, se parsea igual que la
    copia local: misma matriz factura x mes, misma pestana ELISABET.
    """
    return calendario_cobros(ruta_descargada, hoja)


ID_SHEET_ELI = "1EmO9WHz-ewB8objYRnAvoQ2ZBkhnzAbR"
