# -*- coding: utf-8 -*-
"""
============================================================================
 LEASEIR - El Excel de Alejandro, con su flujo de trabajo de verdad
============================================================================
No es un volcado de numeros: es SU libro. La hoja maestra "Forecast Caja -
Mes en Curso" calcula con FORMULAS sobre las subpestanas, exactamente como
hace su fichero 20260717 - Forecast CashFlow.xlsx:

    Forecast   =IFERROR(VLOOKUP($D7,Cobros!$A:$B,2,0),0)
    Adicional  =SUMIFS('Clientes (Nuevos)'!$K:$K,'Clientes (Nuevos)'!$B:$B,$D7)
    Ejecutado  =SUMIF('Cobros Realizados'!$O:$O,$D7,'Cobros Realizados'!$P:$P)
    Pendiente  =Total-Ejecutado
    Proveedor  =-SUMIFS(Proveedores!$S:$S,Proveedores!$F:$F,$D62,
                        Proveedores!$W:$W,"<"&$G$2)

Las columnas clave de las subpestanas estan DONDE EL LAS USA: cliente en B
e importe en K en 'Clientes (Nuevos)'; cliente en O e importe en P en
'Cobros Realizados'; proveedor en F, pendiente en S y mes en W en
'Proveedores'; proveedor en H e importe en L en 'Pagos Realizados'. Asi sus
formulas de siempre funcionan si anade filas o cruza contra su libro.

Todo se rellena desde Holded en cada actualizacion. Si toca una subpestana,
la maestra recalcula sola al abrir: eso es "todo linkado".
============================================================================
"""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AZUL = "1F4D5C"
GRIS = "F1F4F5"
FMT = "#,##0"

f_tit = Font(bold=True, size=14)
f_cab = Font(bold=True, color="FFFFFF")
f_neg = Font(bold=True)
f_gris = Font(color="6B7680", size=9)
r_cab = PatternFill("solid", fgColor=AZUL)
r_sub = PatternFill("solid", fgColor=GRIS)


def _anchos(ws, anchos):
    for j, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = a


def _cab(ws, fila, textos, desde=1):
    for j, t in enumerate(textos, start=desde):
        c = ws.cell(fila, j, t)
        c.font = f_cab
        c.fill = r_cab


def _n(ws, fila, col, v):
    c = ws.cell(fila, col, v)
    if not (isinstance(v, str) and not v.startswith("=")):
        c.number_format = FMT
    return c


def _sub(ws, fila, hasta_col):
    for j in range(1, hasta_col + 1):
        ws.cell(fila, j).fill = r_sub
        ws.cell(fila, j).font = f_neg


# ---------------------------------------------------------------------------
def _hoja_cobros(wb, dc, cobrado_mes_fac, cobrabilidad):
    """
    'Cobros': el pivot por cliente que alimenta el VLOOKUP de la maestra
    (A cliente, B pendiente al inicio del mes) y debajo de las columnas
    de resumen, nada: el detalle factura a factura va en su propia hoja.

    B es el pendiente AL INICIO DEL MES, reconstruido como pendiente de hoy
    mas lo cobrado en el mes. Asi la maestra hace su cuenta de siempre
    (Forecast - Ejecutado = Pendiente) y el pendiente que sale es EXACTO al
    de hoy: la resta no es aproximada, es identidad.
    """
    ws = wb.create_sheet("Cobros")
    _cab(ws, 1, ["Cliente", "Pendiente inicio mes", "Cobrado en el mes",
                 "Pendiente hoy", "Vencido hoy", "Sin vencer",
                 "Cobrable según motor"])
    entra_cli = {}
    if cobrabilidad is not None:
        dfc, _ = cobrabilidad
        if dfc is not None and not dfc.empty:
            for cl, g in dfc.groupby("cliente"):
                entra_cli[cl] = float(g[g["entra"]]["pendiente_cobro"].sum())
    f = 2
    if dc is not None and not dc.empty:
        e = dc.copy()
        e["pend_hoy"] = e["teorico_hoy"] - e["liquidado"]
        e["cob_mes"] = e["num"].map(lambda n: cobrado_mes_fac.get(str(n), 0.0))
        e["venc"] = e["retraso"].clip(lower=0)
        g = e.groupby("cliente").agg(
            pend_hoy=("pend_hoy", "sum"), cob_mes=("cob_mes", "sum"),
            venc=("venc", "sum"), total=("total", "sum"),
            teor=("teorico_hoy", "sum"), liq=("liquidado", "sum")).reset_index()
        g["inicio"] = g["pend_hoy"] + g["cob_mes"]
        g = g[(g["inicio"].abs() > 0.01) | (g["pend_hoy"].abs() > 0.01)]
        for _, r in g.sort_values("inicio", ascending=False).iterrows():
            ws.cell(f, 1, r["cliente"])
            _n(ws, f, 2, round(r["inicio"], 2))
            _n(ws, f, 3, round(r["cob_mes"], 2))
            _n(ws, f, 4, f"=B{f}-C{f}")
            _n(ws, f, 5, round(r["venc"], 2))
            _n(ws, f, 6, round(max(0.0, r["total"] - max(r["teor"], r["liq"])), 2))
            _n(ws, f, 7, round(entra_cli.get(r["cliente"], r["pend_hoy"]), 2))
            f += 1
    _anchos(ws, [42, 18, 16, 14, 13, 13, 13])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(2, f - 1)}"
    return f - 1


def _hoja_cobros_fac(wb, dc, cobrado_mes_fac):
    ws = wb.create_sheet("Cobros (Facturas)")
    _cab(ws, 1, ["Cliente", "Factura", "Vencimiento", "Total",
                 "Debería cobrado (Eli)", "Cobrado total", "Cobrado en el mes",
                 "Vencido", "Sin vencer"])
    f = 2
    if dc is not None and not dc.empty:
        e = dc.copy()
        e["venc"] = e["retraso"].clip(lower=0)
        act = e[(e["teorico_hoy"].abs() > 0.01) | (e["liquidado"].abs() > 0.01)
                | (e["total"].abs() > 0.01)]
        for _, r in act.sort_values(["cliente", "venc"],
                                    ascending=[True, False]).iterrows():
            ws.cell(f, 1, r["cliente"])
            ws.cell(f, 2, str(r["num"]))
            ws.cell(f, 3, str(r.get("vencimiento") or ""))
            _n(ws, f, 4, round(float(r["total"]), 2))
            _n(ws, f, 5, round(float(r["teorico_hoy"]), 2))
            _n(ws, f, 6, round(float(r["liquidado"]), 2))
            _n(ws, f, 7, round(cobrado_mes_fac.get(str(r["num"]), 0.0), 2))
            _n(ws, f, 8, round(float(r["venc"]), 2))
            _n(ws, f, 9, round(max(0.0, float(r["total"])
                                   - max(float(r["teorico_hoy"]),
                                         float(r["liquidado"]))), 2))
            f += 1
    _anchos(ws, [40, 15, 12, 12, 17, 13, 15, 12, 12])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{max(2, f - 1)}"


def _hoja_nuevos(wb, rent, detalle):
    """
    'Clientes (Nuevos)': cliente en B e importe en K, que son las columnas
    que usa su SUMIFS de toda la vida. Rentings del calendario de Eli sin
    factura en Holded, ventas de la fusion LML y ventas aun sin facturar.
    """
    ws = wb.create_sheet("Clientes (Nuevos)")
    ws.cell(1, 1, "Cliente en B e importe en K: las columnas del SUMIFS "
                  "de la maestra. Añade filas y la maestra las recoge."
            ).font = f_gris
    _cab(ws, 2, ["", "Cliente", "Tipo", "Mes", "", "", "", "", "", "",
                 "Importe"], desde=1)
    f = 3
    if rent is not None and not rent.empty:
        for _, r in rent.sort_values("importe", ascending=False).iterrows():
            ws.cell(f, 2, r["cliente"])
            ws.cell(f, 3, "Venta fusión LML" if r.get("tipo") == "fusion"
                    else "Renting sin factura")
            ws.cell(f, 4, str(r["mes"]))
            _n(ws, f, 11, round(float(r["importe"]), 2))
            f += 1
    for x in (detalle or {}).get("sin_facturar") or []:
        ws.cell(f, 2, x.get("concepto", ""))
        ws.cell(f, 3, "Sin facturar")
        _n(ws, f, 11, round(x["unidades"] * x["precio"] * (1 + x["iva"]), 2))
        f += 1
    _anchos(ws, [3, 42, 20, 10, 3, 3, 3, 3, 3, 3, 14])
    ws.freeze_panes = "A3"


def _hoja_realizados(wb, rea, sentido, titulo, col_nombre, col_importe):
    """
    Cobros/Pagos realizados con el tercero y el importe en LAS COLUMNAS QUE
    EL YA USA: O/P en cobros, H/L en pagos. Sus SUMIF siguen funcionando.
    """
    ws = wb.create_sheet(titulo)
    ncols = max(col_importe, col_nombre)
    cab = [""] * ncols
    cab[0:4] = ["Fecha", "Factura", "Banco", "Concepto"]
    cab[col_nombre - 1] = "Cliente" if sentido == "cobro" else "Proveedor"
    cab[col_importe - 1] = "Importe"
    _cab(ws, 1, cab)
    f = 2
    if rea is not None and not rea.empty:
        d = rea[rea["sentido"] == sentido]
        for _, r in d.sort_values("fecha").iterrows():
            ws.cell(f, 1, str(r["fecha"]))
            ws.cell(f, 2, str(r["num"]))
            ws.cell(f, 3, str(r.get("banco") or ""))
            ws.cell(f, 4, str(r.get("concepto") or "")[:80])
            ws.cell(f, col_nombre, r["tercero"])
            _n(ws, f, col_importe, round(float(r["importe"]), 2))
            f += 1
    anchos = [11, 15, 22, 34] + [3] * (ncols - 6) + [40, 13]
    _anchos(ws, anchos[:ncols])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(2, f - 1)}"


def _hoja_proveedores(wb, dp, pagado_mes_fac):
    """
    'Proveedores': proveedor en F, pendiente en S y mes de vencimiento en W,
    las columnas de su -SUMIFS. S es el pendiente AL INICIO DEL MES
    (pendiente hoy + pagado en el mes sobre esa factura), por la misma
    identidad que en Cobros: Forecast - Ejecutado = Pendiente exacto.
    """
    ws = wb.create_sheet("Proveedores")
    cab = [""] * 23
    cab[0:5] = ["Factura", "Fecha venc.", "Tipología", "Estado", ""]
    cab[5] = "Proveedor"                       # F
    cab[18] = "Pendiente inicio mes"           # S
    cab[19] = "Pagado en el mes"               # T
    cab[20] = "Pendiente hoy"                  # U
    cab[22] = "Mes venc. (YYYYMM)"             # W
    _cab(ws, 1, cab)
    f = 2
    if dp is not None and not dp.empty:
        e = dp.copy()
        e["pag_mes"] = e["num"].map(lambda n: pagado_mes_fac.get(str(n), 0.0))
        act = e[(e["pendiente"].abs() > 0.01) | (e["pag_mes"].abs() > 0.01)]
        for _, r in act.sort_values(["proveedor", "vencimiento"]).iterrows():
            v = r.get("vencimiento")
            ws.cell(f, 1, str(r["num"]))
            ws.cell(f, 2, str(v or ""))
            ws.cell(f, 3, str(r.get("tipologia") or ""))
            ws.cell(f, 4, "VENCIDO" if r.get("vencido") else str(r.get("estado") or ""))
            ws.cell(f, 6, r["proveedor"])
            _n(ws, f, 19, round(float(r["pendiente"]) + r["pag_mes"], 2))
            _n(ws, f, 20, round(r["pag_mes"], 2))
            _n(ws, f, 21, f"=S{f}-T{f}")
            if hasattr(v, "year"):
                ws.cell(f, 23, v.year * 100 + v.month)
            f += 1
    _anchos(ws, [15, 12, 14, 12, 3, 40] + [3] * 12 + [18, 14, 13, 3, 16])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:W{max(2, f - 1)}"


def _hoja_bancos(wb, bk, mec):
    ws = wb.create_sheet("Bancos")
    ws.cell(1, 1, "Bottom Up - banco a banco").font = f_tit
    # Inicio y Hoy = contabilidad + lo sin conciliar del extracto (con signo).
    # "Sin contab." dice cuanto de cada saldo es extracto aun sin apuntar.
    _cab(ws, 3, ["Cuenta", "Al empezar el mes", "Hoy", "Variación",
                 "Sin contab.", "Movs"])
    f = 4
    pc = (bk or {}).get("por_cuenta") or []
    for x in pc:
        ws.cell(f, 1, x["cuenta"])
        _n(ws, f, 2, round(x["inicio"], 2))
        _n(ws, f, 3, round(x["hoy"], 2))
        _n(ws, f, 4, f"=C{f}-B{f}")
        _n(ws, f, 5, round(float(x.get("sin_contab") or 0), 2))
        ws.cell(f, 6, x["n"])
        f += 1
    if pc:
        ws.cell(f, 1, "TOTAL")
        for col in "BCDE":
            _n(ws, f, ord(col) - 64, f"=SUM({col}4:{col}{f-1})")
        _sub(ws, f, 6)
        f += 1
    f += 1
    cn = (mec or {}).get("concil") or {}
    _cab(ws, f, ["El levered, contra los bancos", "Importe"]); f += 1
    fila_contable = f
    for nombre, v in [("LEVERED ejecutado = variación contable de bancos",
                       cn.get("contable")),
                      ("Pendiente de conciliar del mes (con signo)",
                       cn.get("pendiente")),
                      ("Descuadre restante", cn.get("descuadre"))]:
        ws.cell(f, 1, nombre)
        _n(ws, f, 2, round(float(v or 0), 2))
        f += 1
    ws.cell(f, 1, "= Variación de saldos bancarios (día 1 → hoy)")
    _n(ws, f, 2, f"=SUM(B{fila_contable}:B{f-1})")
    _sub(ws, f, 2)
    f += 2
    _cab(ws, f, ["Pagos de deuda del mes", "Importe"]); f += 1
    ini_deu = f
    for x in (bk or {}).get("detalle_deuda") or []:
        ws.cell(f, 1, f"{x['cuenta']} {x['nombre'] or 'deuda'}")
        _n(ws, f, 2, round(float(x["importe"]), 2))
        f += 1
    ws.cell(f, 1, "TOTAL deuda")
    _n(ws, f, 2, f"=SUM(B{ini_deu}:B{f-1})" if f > ini_deu else 0)
    _sub(ws, f, 2)
    _anchos(ws, [52, 18, 16, 14, 13, 8])
    return f  # fila del total de deuda


def _hoja_forecast(wb, fc):
    ws = wb.create_sheet("Forecast")
    meses = fc["meses"]
    _cab(ws, 1, ["Concepto"] + [fc["lineas"][m]["etiqueta"] for m in meses])
    CLAVES = [("Cobro clientes", "cobro_clientes"),
              ("Rentings sin factura", "rentings_sin_factura"),
              ("Ventas fusión LML", "ventas_fusion_lml"),
              ("Ventas sin facturar", "ventas_sin_facturar"),
              ("Ajustes", "ajustes_cobros"),
              ("CASH IN", "cash_in"),
              ("Pago proveedores", "pago_proveedores"),
              ("Salarios y SS", "salarios"),
              ("Cuotas S&L", "cuotas_sl"),
              ("Recurrentes", "recurrentes_proyectados"),
              ("Otros fijos", "otros_fijos"),
              ("CASH OUT", "cash_out"),
              ("FCF", "fcf"),
              ("Saldo proyectado", "saldo_proyectado")]
    f = 2
    for nombre, k in CLAVES:
        ws.cell(f, 1, nombre)
        for j, m in enumerate(meses, start=2):
            _n(ws, f, j, round(float(fc["lineas"][m].get(k, 0) or 0), 2))
        if k in ("cash_in", "cash_out", "fcf", "saldo_proyectado"):
            _sub(ws, f, len(meses) + 1)
        f += 1
    _anchos(ws, [26] + [15] * len(meses))


# ---------------------------------------------------------------------------
def _hoja_sl(wb, sl_det):
    """'S&L Cuotas' con la cuota en la columna L, que es donde la busca su
    formula de siempre: =-SUM('S&L Cuotas'!L3:L40)."""
    ws = wb.create_sheet("S&L Cuotas")
    _cab(ws, 2, ["Operación"] + [""] * 10 + ["Cuota"])
    f = 3
    for x in sl_det or []:
        ws.cell(f, 1, x.get("concepto", ""))
        _n(ws, f, 12, round(abs(float(x.get("importe", 0) or 0)), 2))
        f += 1
    _anchos(ws, [46] + [3] * 10 + [13])


def _hoja_visual(wb, etiquetas_cert, ytd):
    """
    'Visual Nacho', identico al suyo: los codigos 1-6 de la columna B de la
    maestra parten el pendiente en seguro / posible / retrasado y la escalera
    C4->C13 acumula desde el unlevered ya ejecutado. C14 es su check.
    """
    M = "'Forecast Caja - Mes en Curso'"
    ws = wb.create_sheet("Visual Nacho")
    ws.cell(2, 2, "RESUMEN").font = f_tit
    ws.cell(4, 2, "Unlevered FCF Hoy")
    _n(ws, 4, 3, f"=+{M}!I109")
    filas = [(5, 1, "Cobros seguros"), (6, 4, "Pagos seguros"),
             (7, None, "Unlevered Seguro"),
             (8, 2, "Cobros posibles"), (9, 5, "Pagos posibles"),
             (10, None, "Unlevered Posible"),
             (11, 3, "Cobros retrasados"), (12, 6, "Pagos retrasados"),
             (13, None, "Unlevered Total")]
    for r, cod, nombre in filas:
        if cod is not None:
            ws.cell(r, 1, cod)
            ws.cell(r, 2, etiquetas_cert.get(cod, nombre))
            _n(ws, r, 3, f"=+SUMIFS({M}!J:J,{M}!B:B,$A{r})")
        else:
            ws.cell(r, 2, nombre)
            base = {7: "C4:C6", 10: "C7:C9", 13: "C10:C12"}[r]
            _n(ws, r, 3, f"=+SUM({base})")
            _sub(ws, r, 3)
    ws.cell(14, 2, "Check")
    _n(ws, 14, 3, f"=+{M}!H109-C13")
    ws.cell(16, 2, "Unlevered meses cerrados (YTD)")
    _n(ws, 16, 3, round(float(ytd or 0), 2))
    ws.cell(17, 2, "Unlevered YTD Seguro"); _n(ws, 17, 3, "=+C16+C7")
    ws.cell(18, 2, "Unlevered YTD Posible"); _n(ws, 18, 3, "=+C16+C10")
    _anchos(ws, [6, 34, 16, 3, 14])


def _hoja_pivot(wb, dc, mapping, norm_fn, meses):
    """
    'Pivot Cobros': los cuadros del Control de cobros, uno al lado de otro
    como en su Pivot Control: pendiente por cliente, DEUDORES (retraso hoy),
    teoricos de los dos meses siguientes, y el corte por comercial (Mapping
    Simple del Control de cobros). Tablas estaticas: se regeneran de Holded
    en cada actualizacion, que es lo que un pivot refrescaria.
    """
    ws = wb.create_sheet("Pivot Cobros")
    if dc is None or dc.empty:
        ws.cell(1, 1, "Sin datos de cobros.")
        return
    e = dc.copy()
    e["pend"] = e["teorico_hoy"] - e["liquidado"]
    e["venc"] = e["retraso"].clip(lower=0)
    m1 = f"teorico_{meses[1]}" if len(meses) > 1 and f"teorico_{meses[1]}" in e.columns else None
    m2 = f"teorico_{meses[2]}" if len(meses) > 2 and f"teorico_{meses[2]}" in e.columns else None

    bloques = [("Pendiente de cobro", "pend", 1),
               ("DEUDORES · retraso a día de hoy", "venc", 4)]
    if m1: bloques.append((f"Teórico {meses[1]}", m1, 7))
    if m2: bloques.append((f"Teórico {meses[2]}", m2, 10))
    for titulo, col, c0 in bloques:
        g = e.groupby("cliente")[col].sum()
        g = g[g.abs() > 0.5].sort_values(ascending=False)
        ws.cell(1, c0, titulo).font = f_neg
        _cab(ws, 2, ["Cliente", "Importe"], desde=c0)
        f = 3
        for cli, v in g.items():
            ws.cell(f, c0, cli); _n(ws, f, c0 + 1, round(float(v), 2)); f += 1
        ws.cell(f, c0, "TOTAL"); _n(ws, f, c0 + 1, round(float(g.sum()), 2))
        _sub(ws, f, 0); ws.cell(f, c0).font = f_neg
        ws.cell(f, c0 + 1).font = f_neg

    # por comercial (Mapping Simple del Control de cobros)
    c0 = 13
    ws.cell(1, c0, "Por comercial (Mapping Simple)").font = f_neg
    if mapping:
        e["comercial"] = e["cliente"].map(
            lambda c: mapping.get(norm_fn(str(c)), "(sin mapping)"))
        g = e.groupby("comercial").agg(pend=("pend", "sum"), venc=("venc", "sum"),
                                       n=("cliente", "nunique"))
        g = g[(g["pend"].abs() > 0.5) | (g["venc"].abs() > 0.5)]
        _cab(ws, 2, ["Comercial", "Pendiente", "Vencido", "Clientes"], desde=c0)
        f = 3
        for com, r in g.sort_values("pend", ascending=False).iterrows():
            ws.cell(f, c0, com)
            _n(ws, f, c0 + 1, round(float(r["pend"]), 2))
            _n(ws, f, c0 + 2, round(float(r["venc"]), 2))
            ws.cell(f, c0 + 3, int(r["n"]))
            f += 1
    else:
        ws.cell(2, c0, "El mapping cliente→comercial no ha llegado: se lee de "
                       "la hoja 'Mapping' del Control de cobros de Eli.")
    _anchos(ws, [40, 13, 2, 40, 13, 2, 40, 13, 2, 40, 13, 2, 28, 13, 12, 9])
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
def generar(ruta, fc, cuadre, meta) -> None:
    """
    La maestra con LAS MISMAS FILAS que su 'Forecast Caja - Mes en Curso':
    Cash IN en la 4, clientes 7-49 y 'Otros' en la 50 absorbiendo el resto
    (como hace el), sin facturar 51-52, ajustes 53-56, Cash OUT 58,
    proveedores 61-101 con 'Variable Mes' en la 102, fijos 103-107,
    Unlevered 109, financiacion 111-113, Levered 115, posicion 117-118,
    bancos 121-128, polizas 130-133 y SU CHECK en K120:

        K120 = I115 - J118   (levered ejecutado - variacion de la posicion)

    La posicion (121-128) es contabilidad MAS lo sin conciliar del extracto,
    asi que K120 da exactamente MENOS el pendiente del mes: K136 (= K120 +
    K135) es el que debe ser ~0. Si no lo es, faltan movimientos que ni el
    diario ni el extracto ven. Sus formulas, sus filas, sus codigos 1-6 en
    la columna B para el Visual Nacho.
    """
    mec = meta.get("mes_en_curso") or {}
    dc = fc.get("detalle_cobros")
    dp = fc.get("detalle_pagos")
    rea = meta.get("realizados")
    rent = fc.get("rentings")
    det = fc.get("detalle") or {}
    eje = fc.get("ejecutado") or {}
    bk = eje.get("banco") or {}
    ban = meta.get("bancos")
    mes0 = fc["meses"][0]
    mes1 = fc["meses"][1] if len(fc["meses"]) > 1 else str(int(mes0) + 1)

    cob_fac, pag_fac = {}, {}
    if rea is not None and not rea.empty:
        for _, r in rea.iterrows():
            k = str(r["num"])
            if r["sentido"] == "cobro":
                cob_fac[k] = cob_fac.get(k, 0.0) + float(r["importe"])
            elif r["sentido"] == "pago":
                pag_fac[k] = pag_fac.get(k, 0.0) + float(r["importe"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Caja - Mes en Curso"

    # ---- subpestanas ------------------------------------------------------
    from fuentes import norm as _norm
    etiquetas_cert = {}
    for lado in ("cobros", "pagos"):
        for k, v in ((meta.get("certidumbre") or {}).get(lado) or {}).items():
            etiquetas_cert[int(k)] = v
    _hoja_visual(wb, etiquetas_cert, meta.get("unlevered_ytd"))
    _hoja_cobros(wb, dc, cob_fac, meta.get("cobrabilidad"))
    _hoja_nuevos(wb, rent[rent["mes"] == mes0] if rent is not None
                 and not rent.empty else rent, {})
    _hoja_realizados(wb, rea, "cobro", "Cobros Realizados", 15, 16)   # O, P
    _hoja_proveedores(wb, dp, pag_fac)
    _hoja_realizados(wb, rea, "pago", "Pagos Realizados", 8, 12)      # H, L
    _hoja_sl(wb, det.get("cuotas_sl"))
    _hoja_pivot(wb, dc, meta.get("mapping_cobros") or {}, _norm, fc["meses"])
    _hoja_cobros_fac(wb, dc, cob_fac)
    _hoja_bancos(wb, bk, mec)
    _hoja_forecast(wb, fc)

    # ---- cabecera ---------------------------------------------------------
    ws.cell(1, 3, "MES (INICIO EN POSICIÓN 2)")
    ws.cell(1, 6, f"{date.today():%d/%m/%Y}")
    ws.cell(2, 3, int(mes0[4:]))
    ws.cell(2, 6, int(mes0)); ws.cell(2, 6).font = f_neg
    ws.cell(2, 8, "Mes sig."); ws.cell(2, 9, int(mes1))
    et = mec.get("etiqueta", mes0).split()[0]
    _cab(ws, 3, ["Mapping Cobros", "Cert.", "", "Concepto", "",
                 f"{et} Forecast", "Adicional", "Total",
                 "Mes Actual (Movs Ejecutados)", "Pendiente Actualizado",
                 "Check", ""])

    def cols(r, F=None, G=None, H=None, I=None, J=None):
        for col, v in zip("FGHIJ", (F, G, H, I, J)):
            if v is not None:
                _n(ws, r, ord(col) - 64, v)

    # ---- Cash IN: clientes 7-49, Otros 50 ---------------------------------
    cert_cli, orden = {}, []
    if dc is not None and not dc.empty:
        e = dc.copy()
        e["pend_hoy"] = e["teorico_hoy"] - e["liquidado"]
        e["cob_mes"] = e["num"].map(lambda n: cob_fac.get(str(n), 0.0))
        e["ini"] = e["pend_hoy"] + e["cob_mes"]
        g = e.groupby("cliente")["ini"].sum().sort_values(ascending=False)
        orden = [c for c in g.index if abs(g[c]) > 0.01]
        if "certidumbre" in e.columns:
            cert_cli = e.groupby("cliente")["certidumbre"].max().to_dict()
    top = orden[:43]

    ws.cell(6, 4, "Cobro Clientes")
    cols(6, *(f"=SUM({c}7:{c}50)" for c in "FGHIJ"))
    _sub(ws, 6, 10)
    for i, cli in enumerate(top):
        r = 7 + i
        ws.cell(r, 1, cli)
        ws.cell(r, 2, int(cert_cli.get(cli, 2)))
        ws.cell(r, 4, cli)
        cols(r,
             f"=IFERROR(VLOOKUP($D{r},Cobros!$A:$B,2,0),0)",
             f"=SUMIFS('Clientes (Nuevos)'!$K:$K,'Clientes (Nuevos)'!$B:$B,$D{r})",
             f"=F{r}+G{r}",
             f"=SUMIF('Cobros Realizados'!$O:$O,$D{r},'Cobros Realizados'!$P:$P)",
             f"=H{r}-I{r}")
    for r in range(7 + len(top), 50):        # filas vacias hasta la 49
        cols(r, 0, 0, f"=F{r}+G{r}", 0, f"=H{r}-I{r}")
    ws.cell(50, 2, 2); ws.cell(50, 4, "Otros")
    cols(50,
         "=SUM(Cobros!$B:$B)-SUM(F7:F49)",
         "=SUM('Clientes (Nuevos)'!$K:$K)-SUM(G7:G49)",
         "=F50+G50",
         "=SUM('Cobros Realizados'!$P:$P)-SUM(I7:I49)",
         "=H50-I50")
    ws.row_dimensions.group(7, 50, outline_level=1)

    sf = det.get("sin_facturar") or []
    for i, r in enumerate((51, 52)):
        if i < len(sf):
            x = sf[i]
            ws.cell(r, 2, 1)
            ws.cell(r, 4, f"{x['concepto']} (Sin facturar)")
            cols(r, round(x["unidades"] * x["precio"] * (1 + x["iva"]), 2),
                 0, f"=F{r}+G{r}", 0, f"=H{r}-I{r}")
        else:
            ws.cell(r, 4, "(Sin facturar)")
            cols(r, 0, 0, f"=F{r}+G{r}", 0, f"=H{r}-I{r}")
    aj = det.get("ajustes") or []
    aj_neg = round(sum(x["importe"] for x in aj if x["importe"] < 0), 2)
    aj_pos = round(sum(x["importe"] for x in aj if x["importe"] > 0), 2)
    ws.cell(53, 2, 3); ws.cell(53, 4, "Ajustes Negativos")
    cols(53, aj_neg, 0, "=F53+G53", 0, "=H53-I53")
    ws.cell(54, 2, 1); ws.cell(54, 4, "Ajustes Positivos (Rentings)")
    ws.cell(54, 12, "los rentings sin factura van por cliente en Adicional; "
                    "aqui irian solo ajustes manuales")
    ws.cell(54, 12).font = f_gris
    cols(54, 0, 0, "=F54+G54", 0, "=H54-I54")
    ws.cell(55, 2, 1); ws.cell(55, 4, "Adicionales Mes")
    cols(55, aj_pos, 0, "=F55+G55", 0, "=H55-I55")
    m0L = fc["lineas"][mes0]
    ws.cell(56, 2, 1); ws.cell(56, 4, "(+) Impuestos")
    cols(56, round(float(m0L.get("impuestos", 0) or 0), 2), 0,
         "=F56+G56", 0, "=H56-I56")

    ws.cell(4, 4, "Cash IN")
    cols(4, *(f"=+{c}6+{c}51+{c}52+{c}53+{c}54+{c}55+{c}56" for c in "FGHIJ"))
    _sub(ws, 4, 10)

    # ---- Cash OUT: proveedores 61-101, Variable 102, fijos 103-107 --------
    orden_p, venc_p = [], {}
    if dp is not None and not dp.empty:
        e = dp.copy()
        e["pag_mes"] = e["num"].map(lambda n: pag_fac.get(str(n), 0.0))
        e["ini"] = e["pendiente"] + e["pag_mes"]
        g = e.groupby("proveedor")["ini"].sum().sort_values(ascending=False)
        orden_p = [p for p in g.index if abs(g[p]) > 0.01]
        venc_p = e.groupby("proveedor")["vencido"].any().to_dict()
    top_p = orden_p[:41]

    ws.cell(60, 4, "Pago Proveedores")
    cols(60, *(f"=SUM({c}61:{c}101)" for c in "FGHIJ"))
    _sub(ws, 60, 10)
    for i, p in enumerate(top_p):
        r = 61 + i
        ws.cell(r, 1, p)
        ws.cell(r, 2, 6 if venc_p.get(p) else 4)
        ws.cell(r, 4, p)
        cols(r,
             f"=-SUMIFS(Proveedores!$S:$S,Proveedores!$F:$F,$D{r},"
             f"Proveedores!$W:$W,\"<\"&$I$2)",
             0, f"=F{r}+G{r}",
             f"=SUMIFS('Pagos Realizados'!$L:$L,'Pagos Realizados'!$H:$H,$D{r})",
             f"=H{r}-I{r}")
    for r in range(61 + len(top_p), 102):
        cols(r, 0, 0, f"=F{r}+G{r}", 0, f"=H{r}-I{r}")
    ws.cell(102, 2, 5); ws.cell(102, 4, "Pago Proveedores Variable Mes")
    cols(102,
         "=-SUMIFS(Proveedores!$S:$S,Proveedores!$W:$W,\"<\"&$I$2)-F60",
         0, "=F102+G102",
         "=SUM('Pagos Realizados'!$L:$L)-SUM(I61:I101)",
         "=H102-I102")
    ws.row_dimensions.group(61, 102, outline_level=1)

    def fijo(r, cod, nombre, total, ejec):
        ws.cell(r, 2, cod); ws.cell(r, 4, nombre)
        cols(r, round(float(total), 2), 0, f"=F{r}+G{r}",
             round(float(ejec), 2), f"=H{r}-I{r}")

    ej_out = {x["concepto"]: x for x in mec.get("cash_out") or []}
    def _co(nombre):
        x = ej_out.get(nombre) or {}
        return float(x.get("ejecutado", 0) or 0), float(x.get("pendiente", 0) or 0)
    sal_e, sal_p = _co("Salarios y seguridad social")
    imp_e, _ = _co("Impuestos")
    sl_e, sl_p = _co("Cuotas sale & leaseback")
    rec_e, rec_p = _co("Recurrentes y otros fijos")
    tar_e, _ = _co("Cuotas de tarjeta ya cargadas")
    otr_e, _ = _co("Otros pagos (comisiones, traspasos, varios)")

    fijo(103, 4, "Salarios", sal_e + sal_p, sal_e)
    ws.cell(104, 2, 4); ws.cell(104, 4, "Cuotas S&L")
    cols(104, "=-SUM('S&L Cuotas'!L3:L40)", 0, "=F104+G104",
         round(sl_e, 2), "=H104-I104")
    if abs(sl_e + sl_p) > 0.005 and not det.get("cuotas_sl"):
        cols(104, round(sl_e + sl_p, 2))
    fijo(105, 4, "Otros (recurrentes, tarjetas, comisiones)",
         rec_e + rec_p + tar_e + otr_e, rec_e + tar_e + otr_e)
    fijo(106, 4, "Adicionales Mes", 0, 0)
    fijo(107, 4, "(-) Impuestos", imp_e, imp_e)

    ws.cell(58, 4, "Cash OUT")
    cols(58, *(f"=+{c}60+{c}102+{c}103+{c}104+{c}105+{c}106+{c}107"
               for c in "FGHIJ"))
    _sub(ws, 58, 10)

    # ---- Unlevered -> financiacion -> Levered (sus filas 109-115) ---------
    ws.cell(109, 4, "Unlevered FCF")
    cols(109, *(f"=+{c}4+{c}58" for c in "FGHIJ"))
    _sub(ws, 109, 10)
    dd = (bk or {}).get("detalle_deuda") or []
    pos_d = round(sum(x["importe"] for x in dd if x["importe"] > 0), 2)
    neg_d = round(sum(x["importe"] for x in dd if x["importe"] < 0), 2)
    fijo(111, 4, "(+) Préstamos (disposiciones)", pos_d, pos_d)
    fijo(112, 4, "(-) Dividendos", 0, 0)
    fijo(113, 4, "(-) Préstamos (principal e intereses)", neg_d, neg_d)
    ws.cell(115, 4, "Levered FCF")
    cols(115, *(f"=+SUM({c}109:{c}113)" for c in "FGHIJ"))
    _sub(ws, 115, 10)

    # ---- posicion bancaria y SU CHECK (117-133) ---------------------------
    ws.cell(117, 4, "Posición bancaria (Sin pólizas)")
    cols(117, "=+F120", None, None, "=+I120")
    _n(ws, 117, 10, "=+I117-F117")
    ws.cell(118, 4, "Posición bancaria (Con pólizas)")
    cols(118, "=+F117+F130", None, None, "=+I117+I130")
    _n(ws, 118, 10, "=+I118-F118")

    ws.cell(120, 4, "LT")
    cols(120, "=SUM(F121:F128)", None, None, "=SUM(I121:I128)")
    _n(ws, 120, 10, "=+I120-F120")
    _n(ws, 120, 11, "=+I115-J118")
    ws.cell(120, 12, "SU CHECK: levered − variación de la posición. Debe ser "
                     "0; si no, mira lo sin conciliar (fila 135).")
    ws.cell(120, 12).font = f_gris
    _sub(ws, 120, 10)
    pc = (bk or {}).get("por_cuenta") or []
    filas_b = pc[:7]
    resto_b = pc[7:]
    for i, x in enumerate(filas_b):
        r = 121 + i
        ws.cell(r, 4, x["cuenta"])
        cols(r, round(x["inicio"], 2), None, None, round(x["hoy"], 2))
        _n(ws, r, 10, f"=+I{r}-F{r}")
    if resto_b:
        r = 121 + len(filas_b)
        ws.cell(r, 4, f"Otras cuentas ({len(resto_b)})")
        cols(r, round(sum(x["inicio"] for x in resto_b), 2), None, None,
             round(sum(x["hoy"] for x in resto_b), 2))
        _n(ws, r, 10, f"=+I{r}-F{r}")
    ws.row_dimensions.group(121, 128, outline_level=1)

    ws.cell(130, 3, "Concedido"); ws.cell(130, 4, "LT Pólizas")
    cols(130, "=SUM(F131:F133)", None, None, "=SUM(I131:I133)")
    _n(ws, 130, 10, "=+I130-F130")
    _sub(ws, 130, 10)
    lim_por_cta = {str(x.get("cuenta")): float(x.get("limite") or 0)
                   for x in meta.get("polizas_limites") or []}
    pol = (ban[ban["tipo"] == "poliza"] if ban is not None and not ban.empty
           else None)
    r = 131
    if pol is not None:
        for _, p in pol.iterrows():
            if r > 133:
                break
            lim = next((v for k, v in lim_por_cta.items()
                        if k and k in str(p["cuenta"])), 0)
            _n(ws, r, 3, lim)
            ws.cell(r, 4, p["cuenta"])
            cols(r, round(float(p["saldo"]), 2), None, None,
                 round(float(p["saldo"]), 2))
            _n(ws, r, 10, f"=+I{r}-F{r}")
            r += 1
    ws.row_dimensions.group(131, 133, outline_level=1)

    # lo que explica el check cuando no es cero
    # El saldo de los bancos (121-128) ya lleva sumado lo sin conciliar, asi
    # que K120 (levered contable - variacion de posicion) da exactamente MENOS
    # el pendiente del mes: K120 + K135 debe ser ~0. Antes se restaba, y con
    # el pendiente dentro de la posicion eso lo contaba dos veces.
    cn = mec.get("concil") or {}
    ws.cell(135, 4, "Sin conciliar del mes (extracto, con signo)")
    _n(ws, 135, 11, round(float(cn.get("pendiente", 0) or 0), 2))
    ws.cell(136, 4, "Check tras lo sin conciliar")
    _n(ws, 136, 11, "=+K120+K135")
    ws.cell(137, 4, "(si esto no es ~0, hay movimientos que faltan: "
                    "díselo al panel)")
    ws.cell(137, 12).font = f_gris
    ws.cell(139, 4, "Saldo hoy"); _n(ws, 139, 6, "=+I117")
    ws.cell(139, 4).font = f_neg
    ws.cell(140, 4, "+ Pendiente del mes (J115)"); _n(ws, 140, 6, "=+J115")
    ws.cell(141, 4, "= Saldo proyectado a cierre"); _n(ws, 141, 6, "=+F139+F140")
    _sub(ws, 141, 6)

    _anchos(ws, [34, 6, 11, 46, 2, 14, 12, 14, 15, 15, 14, 44])
    ws.freeze_panes = "E4"
    ws.sheet_view.zoomScale = 90

    # ---- leeme ------------------------------------------------------------
    ws = wb.create_sheet("Léeme")
    for i, t in enumerate([
        "Cómo funciona este fichero",
        "",
        "Es tu 'Forecast Caja - Mes en Curso', con TUS filas: Cash IN en la 4,",
        "clientes 7-49 y Otros en la 50, sin facturar 51-52, ajustes 53-56,",
        "Cash OUT 58, proveedores 61-101 y Variable Mes en la 102, fijos",
        "103-107, Unlevered 109, financiación 111-113, Levered 115, posición",
        "117-118, bancos 121-128, pólizas 130-133 y tu check en K120",
        "(=I115-J118). Debajo, en K135-K136, lo sin conciliar que lo explica.",
        "",
        "La columna B lleva tus códigos 1-6 y el Visual Nacho los resume igual",
        "que el tuyo. Todas las fórmulas van contra las subpestañas (Cobros,",
        "Clientes (Nuevos) B/K, Cobros Realizados O/P, Proveedores F/S/W,",
        "Pagos Realizados H/L, S&L Cuotas L): si tocas una, la maestra",
        "recalcula al abrir.",
        "",
        "Forecast = pendiente al inicio del mes (pendiente hoy + movido en el",
        "mes), así Total - Ejecutado = pendiente de hoy al euro.",
        "",
        "El fichero se regenera desde Holded en cada actualización: lo que",
        "quieras fijar, a config.yaml (previsiones) o a tu copia local.",
    ], start=1):
        ws.cell(i, 1, t)
        if i == 1:
            ws.cell(i, 1).font = f_tit
    _anchos(ws, [96])

    wb.save(ruta)
